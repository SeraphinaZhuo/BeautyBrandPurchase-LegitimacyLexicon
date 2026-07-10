#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
BPLL Annotation Batch Sampler
==============================
Draws the human-annotation samples from the cleaned Weibo main sample
(METHODS.md §8) and produces:

  annotation_master.csv   <- PI-only: sampled rows WITH all metadata
  trial_batch.xlsx        <- 50 items, identical order for all annotators
  annotator_[A|B|C].xlsx  <- 600 items each (same items, independently
                             shuffled order), text only + empty label columns

Sampling design (all seeds fixed; rerunning reproduces identical output):
  - Posts containing explicit crisis keywords (crisis_flag == 1) are rare
    (~8% of the corpus). A purely random draw of 600 would contain only
    ~48 of them -- too few to evaluate the seven crisis-reaction categories.
  - Therefore: half of each batch is drawn from crisis_flag == 1 rows,
    half from crisis_flag == 0 rows allocated proportionally across
    brand x period cells.
  - Trial batch (25 + 25) is drawn first and excluded from the main pool.
  - Consequence: the annotation sample intentionally over-represents
    crisis vocabulary. It is valid for measuring dictionary accuracy
    (precision / recall), NOT for estimating category prevalence.

Usage:
    python sample_for_annotation.py --data main_sample.csv --outdir annotation/

Command:
    python .\scripts\sample_for_annotation.py --data .\reports\cleaned_weibo\main_sample.csv --outdir .\data\annotation/
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


TRIAL_SEED = 20260709
MAIN_SEED = 20260707

N_TRIAL_CRISIS, N_TRIAL_OTHER = 25, 25
N_MAIN_CRISIS, N_MAIN_OTHER = 350, 350

LABEL_COLS = [
    "价值认同型",
    "社会信号型",
    "安全恐慌",
    "立场抵制",
    "理性辩护",
    "摆烂虚无",
    "割舍挣扎",
    "替代转投",
    "国货红利",
    "主题无关",
    "上下文缺失",
    "无法判断",
    "备注",
]


def proportional_draw(pool: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    """Draw n rows allocated proportionally across brand x period cells."""
    cells = pool.groupby(["brand_category", "time_period"])
    sizes = cells.size()
    alloc = (sizes / sizes.sum() * n).round().astype(int)

    # Rounding can drift by a few rows; correct on the largest cells.
    while alloc.sum() != n:
        alloc[alloc.idxmax()] += 1 if alloc.sum() < n else -1

    parts = [
        cells.get_group(key).sample(n=k, random_state=seed)
        for key, k in alloc.items()
        if k > 0
    ]
    return pd.concat(parts)


def write_annotator_xlsx(items: pd.DataFrame, path: Path, shuffle_seed=None):
    """Write one annotation workbook: 编号 + 文本 + label columns."""
    df = items[["anno_id", "text"]].copy()

    if shuffle_seed is not None:
        df = df.sample(frac=1.0, random_state=shuffle_seed)

    wb = Workbook()
    ws = wb.active
    ws.title = "标注"

    header = ["编号", "发言文本"] + LABEL_COLS
    ws.append(header)

    for _, r in df.iterrows():
        ws.append([r["anno_id"], r["text"]] + [""] * len(LABEL_COLS))

    max_row = len(df) + 1

    # ===== colors: header dark, body light =====
    header_fill_default = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
    header_fill_cd = PatternFill("solid", start_color="E26B0A", end_color="E26B0A")      # C:D 深橙
    header_fill_ek = PatternFill("solid", start_color="76933C", end_color="76933C")      # E:K 深绿
    header_fill_ln = PatternFill("solid", start_color="366092", end_color="366092")      # L:N 深蓝

    body_fill_cd = PatternFill("solid", start_color="FDE9D9", end_color="FDE9D9")         # C:D 浅橙
    body_fill_ek = PatternFill("solid", start_color="EBF1DE", end_color="EBF1DE")         # E:K 浅绿
    body_fill_ln = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")         # L:N 浅蓝

    # ===== header style =====
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if 3 <= c <= 4:        # C:D
            cell.fill = header_fill_cd
            cell.font = Font(bold=True, color="FFFFFF")
        elif 5 <= c <= 11:     # E:K
            cell.fill = header_fill_ek
            cell.font = Font(bold=True, color="FFFFFF")
        elif 12 <= c <= 14:    # L:N
            cell.fill = header_fill_ln
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.fill = header_fill_default
            cell.font = Font(bold=True)

    # ===== body group fills =====
    for col in "CD":
        for row in range(2, max_row + 1):
            ws[f"{col}{row}"].fill = body_fill_cd

    for col in "EFGHIJK":
        for row in range(2, max_row + 1):
            ws[f"{col}{row}"].fill = body_fill_ek

    for col in "LMN":
        for row in range(2, max_row + 1):
            ws[f"{col}{row}"].fill = body_fill_ln

    # ===== column widths =====
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 80

    for col in "CDEFGHIJKLMN":
        ws.column_dimensions[col].width = 12

    ws.column_dimensions["O"].width = 30

    # ===== alignments =====
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=14):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=15, max_col=15):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ===== data validation: C:N only, O is remark =====
    dv = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
    dv.error = "请只选择空白或 ✓。"
    dv.errorTitle = "无效输入"
    dv.prompt = "选择 ✓ 表示该标签适用；留空表示不适用。"
    dv.promptTitle = "标签勾选"

    ws.add_data_validation(dv)
    dv.add(f"C2:N{max_row}")

    ws.freeze_panes = "C2"

    wb.save(path)


def run(data_path: str, outdir: str):
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(data_path, encoding="utf-8-sig")

    text_col = "text_norm" if "text_norm" in df.columns else "text_content"
    df = df.rename(columns={text_col: "text"})

    crisis = df[df["crisis_flag"] == 1]
    other = df[df["crisis_flag"] == 0]

    trial = pd.concat(
        [
            crisis.sample(n=N_TRIAL_CRISIS, random_state=TRIAL_SEED),
            proportional_draw(other, N_TRIAL_OTHER, TRIAL_SEED),
        ]
    )

    remaining_crisis = crisis.drop(trial.index, errors="ignore")
    remaining_other = other.drop(trial.index, errors="ignore")

    main = pd.concat(
        [
            remaining_crisis.sample(n=N_MAIN_CRISIS, random_state=MAIN_SEED),
            proportional_draw(remaining_other, N_MAIN_OTHER, MAIN_SEED),
        ]
    )

    trial = trial.copy()
    main = main.copy()

    trial["batch"] = "trial"
    main["batch"] = "main"

    master = pd.concat([trial, main]).reset_index(drop=True)
    master["anno_id"] = (
        ["T%03d" % i for i in range(1, len(trial) + 1)]
        + ["M%03d" % i for i in range(1, len(main) + 1)]
    )

    master.to_csv(out / "annotation_master.csv", index=False, encoding="utf-8-sig")

    write_annotator_xlsx(
        master[master["batch"] == "trial"],
        out / "trial_batch.xlsx",
        shuffle_seed=TRIAL_SEED + 999,
    )

    for name, shuffle_seed in (("A", 101), ("B", 202), ("C", 303)):
        write_annotator_xlsx(
            master[master["batch"] == "main"],
            out / f"annotator_{name}.xlsx",
            shuffle_seed=shuffle_seed,
        )

    print(
        f"trial={len(trial)}  main={len(main)}  "
        f"main_crisis={int(main['crisis_flag'].sum())}"
    )
    print("outputs:", sorted(p.name for p in out.iterdir()))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True)
    ap.add_argument("--outdir", default="annotation")
    args = ap.parse_args()
    run(args.data, args.outdir)